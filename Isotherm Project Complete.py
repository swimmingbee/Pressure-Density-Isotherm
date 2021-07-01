

# import applicalble libraries

import numpy as np
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt

# assign directory variable as a string of desired directory
Directory = '/Users/christiaanroelofse/Documents/GitHub/Isotherm_Project_Template.xlsx'

# load data from excel file (parameters filled in on excel file)
wb = openpyxl.load_workbook(str(Directory))
df = pd.read_excel(str(Directory))

# check loaded data
# df.head()
# df.shape

# assigning variables
sheet = wb["Sheet1"]
T_start_celsius = sheet['B2'].value
T_start = T_start_celsius + 273
delta_T = sheet['B3'].value
isotherms = sheet['B4'].value
R = sheet['B5'].value
P_start = sheet['D2'].value
P_stop = sheet['D3'].value
delta_P = sheet['D4'].value

# specific molecule data
molecule = sheet['B7'].value
Tc_celsius = sheet['B8'].value
Tc = Tc_celsius + 273
Pc = sheet['B9'].value
w = sheet['B10'].value
mw = sheet['B11'].value


# check if parameters are correct
print('\nMolecule: ', molecule,
    '\nCritical temperature: ', Tc, 'K',
    '\nCritical pressure: ', Pc, 'bar',
    '\nAccentric factor: ', w,
    '\nMolecular weight: ', mw, 'g/mol'
    '\nStarting temperature: ', T_start, 'K'
    '\nTemperature increments: ', delta_T,
    '\nNumber of isotherms: ', isotherms,
    '\nGas constant: ', R, 'cm^3*bar/mol*K'
    '\nStarting pressure: ', P_start, 'bar'
    '\nStopping pressure: ', P_stop,'bar'
    '\nPressure increments: ', delta_P, 'bar'
    '\n'
    )


# arrays to store calculated values
array1 = []
array2 = []
array3 = []
arrayT = []
arrayP = []

# pressure values being tested
for i in range(P_start, P_stop+1, delta_P):
    arrayP.append(i)

# temperature values being tested
for t in range(T_start, isotherms*delta_T+T_start, delta_T):
    arrayT.append(t)

# loop for which values get stored in what array
def array_maker(n):
    if Tcount == 1:
        array1.append(n)
    elif Tcount == 2:
        array2.append(n)
    else:
        array3.append(n)



# calculations - all molecule specific values, coefficients, variables, and densities at all temperatures and pressures
def tloop():
    global Tcount
    Tcount = 0
    # arrays to store variables to check if errors occur
    listTr = []
    listalpha = []
    listA = []
    for t in range(T_start, isotherms*delta_T+T_start, delta_T):
        Tr = t/Tc
        listTr.append(Tr)
        alpha = (1 + (0.37464 + 1.54226*w - 0.26992*w**2)*(1-Tr**0.5))**2
        listalpha.append(alpha)
        B = (0.07780*R*Tc)/Pc
        A = (0.45724 * ((R**2)*(Tc**2)/Pc)) * alpha
        listA.append(A)
        Tcount += 1
        def ploop():
            Pcount = 0
            # arrays to store variables to check if errors occur
            listAA = []
            listBB = []
            for p in range(P_start, P_stop+1, delta_P):
                AA = (A*p)/((R**2)*(t**2))
                listAA.append(AA)
                BB = (B*p)/(R*t)
                listBB.append(BB)
                RA1 = 1
                RA2 = BB-1
                RA3 = AA-2*BB-3*BB**2
                RA4 = BB**3+BB**2-AA*BB
                # if unexplainable errors occur, change NNN = 0 to NNN = 1
                NNN = 0
                #or NNN = 1
                Pcount += 1
                def FindDen(RA1,RA2,RA3,RA4,NNN):
                    A1 = ((3 * RA3)-(RA2**2))/3
                    B1 = ((2*RA2**3)-(9*RA2 * RA3) + (27*RA4))/27
                    terms = [RA1, RA2, RA3, RA4]
                    Z = np.roots(terms)
                    if NNN == 0:
                            Z = max(np.roots(terms))
                    if NNN == 1:
                            Z = min(np.roots(terms))
                    global den
                    den = (p*mw)/(R*t*Z)
                    array_maker(float(den))
                    return(float(den))
                FindDen(RA1,RA2,RA3,RA4,NNN)
        ploop()
tloop()

# array of arrays of densities calculated at each pressure and each temperature - used for plot below
arrayDen = []
arrayDen.append(array1)
arrayDen.append(array2)
arrayDen.append(array3)


# plot results in pyplot
axes = plt.axes()
def plot_results(dens):
    tac = 0
    for i in dens:
        plt.plot(i, arrayP, label=(str(arrayT[tac]) + ' K'))
        tac += 1
    plt.legend(loc=2)
    plt.xlabel(str(molecule)+' Density (g/cc)')
    plt.ylabel('Pressure (bar)')
    plt.title('Pressure-Density Isotherm of '+str(molecule))
    axes.grid()
    plt.show()
plot_results(arrayDen)


# creating a dataframe of the calculated data 
data = pd.DataFrame(np.transpose(arrayDen))
data.insert(loc=0, column='Pressure (bar)', value=arrayP)
col_names = ['Pressure (bar)', 'Density (g/cc) at 25 C', ' Density ( g/cc) at 50 C', 'Density (g/cc) at 75 C', ]
data.columns = col_names
print('\n Table of Ethylene densities at specified temperatures and pressures:\n')
print(data)
print('\n')


# send results to excel
# customizable to cater to spreadsheet
def to_XL():
    # temp for 25 degrees
    sheet.cell(row = 16, column = 1).value = T_start_celsius
    # temp for 50 degrees
    sheet.cell(row = 16, column = 4).value = T_start_celsius + delta_T
    # temp for 75 degrees
    sheet.cell(row = 16, column = 7).value = T_start_celsius + 2*delta_T
    
    # pressure columns
    for i in range(0, len(arrayP)):
        sheet.cell(row = i+16, column = 2).value = arrayP[i]
        sheet.cell(row = i+16, column = 5).value = arrayP[i]
        sheet.cell(row = i+16, column = 8).value = arrayP[i]

    # densities calculated and sent
    for i in range(0,len(array1)):
        sheet.cell(row = i+16, column = 3).value = array1[i]
    for i in range(0,len(array2)):
        sheet.cell(row = i+16, column = 6).value = array2[i]
    for i in range(0,len(array3)):
        sheet.cell(row = i+16, column = 9).value = array3[i]

# un-comment two lines below to run 'to_XL' function and save the spreadsheet as a new file
to_XL()
wb.save(filename='Isotherm Project - Results.xlsx')